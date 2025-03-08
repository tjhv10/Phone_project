import logging
import os
import random
import re
import subprocess
import threading
from time import sleep
import cv2
import easyocr
from fuzzywuzzy import fuzz
import numpy as np
import requests
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from PIL import Image
from common_area_items import *
import pytesseract
from PIL import ImageEnhance
from PIL import ImageFilter
import uiautomator2 as u2
import pandas as pd
from openpyxl import load_workbook
import shutil


keyboard_dic = {
    "q": [(40, 1200)],
    "w": [(110, 1200)],
    "e": [(180, 1200)],
    "r": [(250, 1200)],
    "t": [(320, 1200)],
    "y": [(390, 1200)],
    "u": [(460, 1200)],
    "i": [(530, 1200)],
    "o": [(600, 1200)],
    "p": [(670, 1200)],
    "a": [(70, 1285)],
    "s": [(140, 1285)],
    "d": [(210, 1285)],
    "f": [(280, 1285)],
    "g": [(350, 1285)],
    "h": [(420, 1285)],
    "j": [(490, 1285)],
    "k": [(560, 1285)],
    "l": [(630, 1285)],
    "z": [(150, 1400)],
    "x": [(220, 1400)],
    "c": [(290, 1400)],
    "v": [(360, 1400)],
    "b": [(430, 1400)],
    "n": [(500, 1400)],
    "m": [(570, 1400)],
    ".": [(570,1500)],
    ",": [(150,1500)],
    " ": [(400,1500)],
    "_": [(65,1500),(250,1500),(65,1500)],
    "0":[(65,1500),(680,1190),(65,1500)],
    "1":[(65,1500),(40,1190),(65,1500)],
    "2":[(65,1500),(110,1190),(65,1500)],
    "3":[(65,1500),(180,1190),(65,1500)],
    "4":[(65,1500),(250,1190),(65,1500)],
    "5":[(65,1500),(320,1190),(65,1500)],
    "6":[(65,1500),(390,1190),(65,1500)],
    "7":[(65,1500),(460,1190),(65,1500)],
    "8":[(65,1500),(530,1190),(65,1500)],
    "9":[(65,1500),(610,1190),(65,1500)],
    "'":[(65,1500),(288,1400),(65,1500)],
    "@":[(65,1500),(72,1275),(65,1500)],
    "#":[(65,1500),(145,1275),(65,1500)],
    "$":[(65,1500),(219,1275),(65,1500)],
    "%":[(65,1500),(292,1275),(65,1500)],
    "&":[(65,1500),(360,1275),(65,1500)],
    "!":[(65,1500),(502,1387),(65,1500)],
    "?":[(65,1500),(578,1387),(65,1500)],    
}
keyboard_dic_only_nums = {
    "1":[(72,1160)],
    "2":[(280,1160)],
    "3":[(475,1160)],
    "4":[(72,1265)],
    "5":[(280,1265)],
    "6":[(475,1265)],
    "7":[(72,1380)],
    "8":[(280,1380)],
    "9":[(475,1380)],
    "0":[(280,1500)],
}


def type_keyboard(d, text:str, keyboard = keyboard_dic):
    """
    Simulates tapping on the screen using the keyboard coordinates for each character in the text.
    """
    is_upper = False
    for char in text:
        if char.isupper():
            is_upper = True
            char = char.lower()
        if char not in keyboard and not char.isupper():
            char = " "  
        for i,_ in enumerate(keyboard[char]):
            if is_upper:
                is_upper = False
                d.click(50,1386)
                sleep(0.2)
            x, y = keyboard[char][i]
            d.click(x, y)  # Simulate a tap on the screen at the corresponding coordinates
            update_results_file("Actions")
            sleep(random.uniform(0.1, 0.2))  # Add a small delay between taps
                

def search_sentence(d, name: str, plat, tolerance=20, usegpu=True, y_min=0, y_max=1650, bestMatch=False, min_word_length=0,screen_shot=None,return_always=False):
    """
    Searches for a word or sentence in the screenshot using OCR.

    Args:
        d: Device object.
        name (str): The text to search for (word or sentence).
        plat: Platform identifier.
        tolerance (int): Allowed similarity percentage (default=20).
        usegpu (bool): Whether to use GPU for OCR (default=True).
        y_min (int): Minimum Y-coordinate boundary (default=0).
        y_max (int): Maximum Y-coordinate boundary (default=infinity).
        min_word_length (int): Minimum length of the word to be returned (default=2).
    
    Returns:
        tuple: Coordinates of the match center (x, y) or None if no match found.
    """
    if screen_shot is None:
        screen_shot = take_screenshot(d, threading.current_thread().name, plat)
    logging.info(f"{threading.current_thread().name}:{d.serial} Searching for text: {name}")

    while plat == "twi" and name.lower().strip() == "israel":
        name = random.choice(twitter_handles)
        logging.info(f"{threading.current_thread().name}:{d.serial} Searching for text: {name}")
    
    # Determine if we're searching for a word or a sentence
    is_word_search = len(name.split()) == 1 and name.__len__() < 30

    # Initialize the OCR reader
    reader = easyocr.Reader(['en'], gpu=usegpu)

    # Perform OCR
    result = reader.readtext(screen_shot, detail=1)

    best_similarity = 0  # Initialize with the lowest possible score (0%)
    best_match_text = ""  # To store the closest matching text (word or sentence)
    best_match_bbox = []  # To store the bounding box for the best match

    # Process the name for matching
    processed_name = name.strip()

    if is_word_search:
        # Word search logic
        for detection in result:
            bbox, text, _ = detection
            top_left, _, bottom_right, _ = bbox

            # Filter out text outside the vertical boundaries
            if top_left[1] < y_min or bottom_right[1] > y_max:
                continue

            # Split the detected text into words
            words = text.strip().split()

            for word in words:
                # Check if the word is long enough before comparing
                if len(word) < min_word_length:
                    continue

                similarity_score = fuzz.ratio(processed_name.lower(), word.lower())
                
                if similarity_score > best_similarity:
                    best_similarity = similarity_score
                    best_match_text = word
                    best_match_bbox = bbox
    else:
        # Sentence search logic
        # Group detected text into rows by their y-coordinates
        rows = {}
        for detection in result:
            bbox, text, _ = detection
            top_left, _, bottom_right, _ = bbox
            y_center = (top_left[1] + bottom_right[1]) // 2

            # Filter out text outside the vertical boundaries
            if top_left[1] < y_min or bottom_right[1] > y_max:
                continue

            # Skip rows with single-character text
            if len(text.strip()) == 1:
                continue

            # Group rows by y-coordinate
            row_key = round(y_center, -1)
            if row_key not in rows:
                rows[row_key] = []
            rows[row_key].append((bbox, text.strip()))

        # Combine rows into multi-line text blocks
        combined_blocks = []
        sorted_row_keys = sorted(rows.keys())
        for i, row_key in enumerate(sorted_row_keys):
            current_block = " ".join(text for _, text in rows[row_key])
            combined_bboxes = [bbox for bbox, _ in rows[row_key]]

            for offset in range(1, 3):
                if i + offset < len(sorted_row_keys):
                    next_row_key = sorted_row_keys[i + offset]
                    if abs(next_row_key - row_key) <= 0:
                        next_row_text = " ".join(text for _, text in rows[next_row_key])
                        current_block += " " + next_row_text
                        combined_bboxes.extend(bbox for bbox, _ in rows[next_row_key])
                    else:
                        break

            combined_blocks.append((current_block, combined_bboxes))

        # Search for the best match across all combined blocks
        for combined_text, bboxes in combined_blocks:
            if "Go to" in combined_text:
                combined_text = combined_text.replace("Go to", '')

            similarity_score = fuzz.ratio(processed_name, combined_text)

            if similarity_score > best_similarity:
                best_similarity = similarity_score
                best_match_text = combined_text
                best_match_bbox = bboxes

    if best_similarity >= (100 - tolerance) and best_match_bbox:
        # Calculate the center of the bounding box for the best match
        if is_word_search:
            top_left_x, top_left_y = best_match_bbox[0]
            bottom_right_x, bottom_right_y = best_match_bbox[2]
        else:
            all_top_lefts = [bbox[0] for bbox in best_match_bbox]
            all_bottom_rights = [bbox[2] for bbox in best_match_bbox]

            top_left_x = min(coord[0] for coord in all_top_lefts)
            top_left_y = min(coord[1] for coord in all_top_lefts)
            bottom_right_x = max(coord[0] for coord in all_bottom_rights)
            bottom_right_y = max(coord[1] for coord in all_bottom_rights)

        center_x = (top_left_x + bottom_right_x) // 2
        center_y = (top_left_y + bottom_right_y) // 2

        logging.info(f"{threading.current_thread().name}:{d.serial} Best match found: \"{best_match_text}\" with similarity: {best_similarity}% in {center_x,center_y}")
        if bestMatch:
            return best_match_text
        return int(center_x), int(center_y)

    # Log the best match even if it doesn't meet the tolerance
    if best_match_text:
        logging.info(f"{threading.current_thread().name}:{d.serial} Closest match: \"{best_match_text}\" with similarity: {best_similarity}%")
    if return_always:
        return best_match_text

    logging.info(f"{threading.current_thread().name}:{d.serial} No sufficiently similar text was found.")
    return None


def take_screenshot(d, thread=threading.current_thread().name, app="twi", crop_area=None):
    """
    Takes a screenshot of the device and optionally crops it to a specific region.

    Args:
        d: Device object.
        thread (str): Thread name for logging and filename.
        app (str): Application name for filename.
        crop_area (tuple): Optional tuple defining the cropping area (x1, y1, x2, y2).

    Returns:
        str: Path to the saved screenshot or cropped image.
    """
    filename = f"Screenshots/{thread}-screenshot_{app}.png"
    cropped_filename = f"Screenshots/{thread}-screenshot_{app}_cropped.png"

    logging.info(f"{thread}:{d.serial} Taking screenshot...")
    d.screenshot(filename)
    logging.info(f"Screenshot saved as {filename}.")

    # Crop the image if crop_area is specified
    if crop_area:
        x1, y1, x2, y2 = crop_area
        with Image.open(filename) as img:
            cropped_img = img.crop((x1, y1, x2, y2))  # Crop the specified region
            cropped_img.save(cropped_filename)
            logging.info(f"Cropped screenshot saved as {cropped_filename}.")
        return cropped_filename

    return filename


def find_best_match(image_path, users_template_path, d):
    """
    Finds the best match of a user's button icon in the screenshot using template matching.
    """
    sleep(0.5)
    logging.info(f"{threading.current_thread().name}:{d.serial} Starting find_best_match function")
    
    img = cv2.imread(image_path)
    template = cv2.imread(users_template_path)

    if img is None or template is None:
        logging.info(f"{threading.current_thread().name}:{d.serial} Error loading images.")
        return None

    h, w = template.shape[:2]
    result = cv2.matchTemplate(img, template, cv2.TM_CCOEFF_NORMED)
    threshold = 0.8
    loc = np.where(result >= threshold)

    matches = []
    for pt in zip(*loc[::-1]):
        matches.append((pt, result[pt[1], pt[0]]))

    if matches:
        # Get the best match (highest confidence value)
        best_match = max(matches, key=lambda x: x[1])
        best_coordinates = (best_match[0][0] + w // 2, best_match[0][1] + h // 2)
        best_value = best_match[1]
        logging.info(f"{threading.current_thread().name}:{d.serial} Best match found with value: {best_value} at {best_coordinates}")
    else:
        # If no matches found above threshold, find the closest match
        _, max_val, _, max_loc = cv2.minMaxLoc(result)
        best_coordinates = (max_loc[0] + w // 2, max_loc[1] + h // 2)
        best_value = max_val
        logging.info(f"{threading.current_thread().name}:{d.serial} No matches above threshold, closest match found with value: {best_value} at {best_coordinates}")
        return None
    
    logging.info(f"{threading.current_thread().name}:{d.serial} Finished find_best_match function")
    
    return best_coordinates

def handle_user_selection(d,report_dict):
    logging.info("Select a report reason:")
    numbered_report_dict = show_tree(report_dict)

    # User input for selection
    user_choice = input("Enter the number of the report reason you want to select: ")

    if user_choice.isdigit() and int(user_choice) in numbered_report_dict:
        action = numbered_report_dict[int(user_choice)]
        if isinstance(action, dict):  # If the selection has subcategories
            handle_user_selection(action)  # Show subcategories
        else:
            execute_action(d,action,report_dict)  # Execute the action for the selected reason
    else:
        logging.info("Invalid selection. Please enter a valid number.")

def show_tree(report_dict, level=0):
    numbered_dict = {}
    count = 1
    for key in report_dict.keys():
        logging.info("  " * level + f"{count}. {key}")
        numbered_dict[count] = key  # Store the original key for action retrieval
        count += 1
        if isinstance(report_dict[key], dict):
            # Recursive call for subcategories
            sub_count = show_tree(report_dict[key], level + 1)
            numbered_dict.update(sub_count)
    return numbered_dict

def execute_action(d,reason,report_dict):
    # Execute the corresponding action for the selected reason
    if reason in report_dict:
        action = report_dict[reason]
        actions = action.split(':')
        logging.info(f"Executing action for '{reason}': {actions}")
        sleep(2)
        for act in actions:
            exec(act)
            sleep(5)  
    else:
        logging.info("No action found for this reason.")
        
        

file_lock = threading.Lock()
file_path = "result.txt"
backup_path = "result_backup.txt"

# Expected format (order and keys must match)
EXPECTED_KEYS = [
    "Likes", "Comments", "Follows", "Reposts", "Posts reported", "Accounts reported", "Actions"
]

def is_valid_format(file_path):
    """Checks if the file follows the expected format."""
    try:
        with open(file_path, "r") as file:
            lines = file.readlines()

        if len(lines) != len(EXPECTED_KEYS):
            return False

        parsed_keys = []
        for line in lines:
            match = re.match(r"^(.+) - (\d+)$", line.strip())
            if not match:
                return False
            key, value = match.groups()
            if key not in EXPECTED_KEYS:
                return False
            parsed_keys.append(key)

        return parsed_keys == EXPECTED_KEYS  # Ensure order is correct

    except Exception:
        return False  # If any error occurs, assume invalid format



def update_results_file(action_type, counter=1):
    """
    Safely updates the results file with the incremented count for the given action.
    Creates a backup before modifying the file if the backup file follows the correct format.
    
    Parameters:
    action_type (str): The action type to update ('Likes', 'Comments', 'Follows', 'Reposts',
                       'Posts reported', 'Accounts reported', 'Actions').
    counter (int): The amount to increment the count by (default is 1).
    """
    print(f"Updating file for action: {action_type}")
    with file_lock:  # Ensure only one thread accesses the file at a time
        try:
            # Create a backup only if the file follows the correct format
            if os.path.exists(file_path) and is_valid_format(file_path):
                shutil.copy(file_path, backup_path)

            # Load current values
            stats = {key: 0 for key in EXPECTED_KEYS}  # Ensure all keys exist
            if os.path.exists(file_path):
                with open(file_path, "r") as file:
                    for line in file:
                        key, value = line.strip().split(" - ")
                        if key in stats:
                            stats[key] = int(value)

            # Increment the relevant action count
            if action_type in stats:
                stats[action_type] += counter
            else:
                stats[action_type] = counter  # Initialize if not present

            # Update the total "Actions" count
            stats["Actions"] = sum(v for k, v in stats.items() if k != "Actions")

            # Write updated values to a temp file and then rename it (atomic write)
            temp_file = file_path + ".tmp"
            with open(temp_file, "w") as file:
                for key in EXPECTED_KEYS:  # Maintain order
                    file.write(f"{key} - {stats[key]}\n")

            # Replace original file with the updated file
            os.replace(temp_file, file_path)

        except Exception as e:
            print(f"Error updating file: {e}")
            # Restore from backup if something goes wrong
            if os.path.exists(backup_path):
                shutil.copy(backup_path, file_path)
                print("Restored from backup.")



def startAccount():
    # Set up the WebDriver to connect to the existing Chrome instance with remote debugging
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("debuggerAddress", "localhost:9222")  # Connect to the remote debugging port

    # Set up the Chrome WebDriver path
    chrome_driver_path = 'C:/Users/goldf/OneDrive/Documents/chromedriver_win32/chromedriver.exe'

    # Initialize the WebDriver with the specified options to connect to the existing browser
    driver = webdriver.Chrome(options=chrome_options)

    # Open ChatGPT in the existing Chrome window (this will use the existing session)
    driver.get("https://chat.openai.com/")

    # Wait for the page to load
    time.sleep(15)

    # Send a prompt
    prompt_text = "Generate an image of a futuristic city skyline at sunset with flying cars."
    chat_input = driver.find_element(By.XPATH, '//*[@id="prompt-textarea"]/p')  # Find the text input box
    chat_input.send_keys(prompt_text)
    time.sleep(1)
    chat_input.send_keys(Keys.ENTER)

    # Wait for the image to be generated (adjust timing as needed)
    time.sleep(30)

    # Locate the generated image and download it
    try:
        image_element = driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/main/div[1]/div[1]/div/div/div/div/article[2]/div/div/div[2]/div/div[1]/div[1]/div/div/div/div[2]/img')
        image_url = image_element.get_attribute("src")

        # Download the image using requests
        image_response = requests.get(image_url)
        
        # Save the image
        if image_response.status_code == 200:
            with open("generated_image.png", 'wb') as f:
                f.write(image_response.content)
            logging.info("Image downloaded successfully.")
        else:
            logging.info("Failed to download the image.")

    except Exception as e:
        logging.info(f"Error: {e}")

    # Close the browser if necessary
    driver.quit()


def reopen_app(d, package_name, wait_time=5):
    """
    Stops and reopens an app on the device.

    Parameters:
    d (uiautomator2.Device): The connected device instance.
    package_name (str): The package name of the app to be reopened (e.g., "com.twitter.android").
    wait_time (int): Time in seconds to wait after reopening the app.
    """
    try:
        # Check if the app is running
        if package_name in d.app_list_running():
            logging.info(f"{threading.current_thread().name}:{d.serial} Stopping app: {package_name}")
            d.app_stop(package_name)  # Stop the app
            sleep(2)

        # Start the app
        logging.info(f"{threading.current_thread().name}:{d.serial} Starting app: {package_name}")
        d.app_start(package_name)  # Start the app

        # Wait for the app to fully load
        sleep(wait_time)

        # Confirm the app is running
        if package_name in d.app_list_running():
            logging.info(f"{threading.current_thread().name}:{d.serial} Successfully reopened app: {package_name}")
        else:
            logging.warning(f"{threading.current_thread().name}:{d.serial} Failed to reopen app: {package_name}")

    except Exception as e:
        logging.error(f"Error while reopening app {package_name}: {e}")

def is_app_installed(d, package_name):
    """
    Checks if a specific app is installed on the device.

    Parameters:
    d (uiautomator2.Device): The connected device instance.
    package_name (str): The package name of the app to check (e.g., "com.twitter.android").

    Returns:
    bool: True if the app is installed, False otherwise.
    """
    return package_name in d.app_list()
    

def open_vpn(d):
    if not is_app_installed(d, "com.nordvpn.android"):
        return
    logging.info(f"{threading.current_thread().name}: {d.wlan_ip} : Opened nordVPN!")
    d.app_start("com.nordvpn.android")
    sleep(15)
    d.click(400, 200)
    sleep(5)
    count = 0
    while not search_sentence(d, "Pause", "twi"):
        if count == 1:
            try:
                d.click(*search_sentence(d, "Quick connect", "vpn", tolerance=30))
            except:
                logging.info(f"{threading.current_thread().name}: {d.wlan_ip} : Maybe offline or somthing.")
                return
        if count == 2:
            logging.info(f"{threading.current_thread().name}: {d.wlan_ip} : Couldn't find the pause button.")
            return
        count += 1
        logging.info(f"{threading.current_thread().name}: {d.wlan_ip} : Trying to reconnect...")
        sleep(100)
        if not d.info['screenOn']:
            count = 0
            d.shell("input keyevent KEYCODE_POWER")
        sleep(10)


def close_apps(device):
    device.app_stop("com.twitter.android")
    sleep(0.5)
    device.app_stop("com.zhiliaoapp.musically")
    sleep(0.5)
    device.app_stop("com.nordvpn.android")
    sleep(0.5)
    device.app_stop("com.instagram.lite")
    sleep(0.5)
    device.app_stop("com.google.android.gm")
    logging.info(f"{device.wlan_ip} closed apps.")


def advanced_preprocess_image_inplace(image_path):
    """
    Advanced preprocessing to enhance OCR performance and saves it in the same path.

    Args:
        image_path (str): Path to the input image.

    Returns:
        None
    """
    # Load the image
    image = Image.open(image_path)

    # Convert to grayscale
    image = image.convert("L")

    # Enhance sharpness
    enhancer = ImageEnhance.Sharpness(image)
    image = enhancer.enhance(2.0)  # Increase sharpness

    # Apply binarization (adaptive thresholding for better contrast)
    image = image.point(lambda x: 0 if x < 150 else 255)

    # Resize the image to improve OCR accuracy
    image = image.resize((image.width * 3, image.height * 3), Image.Resampling.LANCZOS)

    # Apply more aggressive noise reduction
    image = image.filter(ImageFilter.MedianFilter(size=5))

    # Save the preprocessed image back to the same path
    image.save(image_path)
    print(f"Advanced preprocessed image saved in place at {image_path}")


def image_to_string(image_path,number = True):
    """
    Converts text in an image to a string using OCR.

    Args:
        image_path (str): Path to the image file.

    Returns:
        str: Cleaned and processed extracted text from the image.
    """
    print(f"Extracting text from image: {image_path}")
    try:
        # Preprocess the image in place for better OCR performance
        
        # Extract text using pytesseract
        if number:
            extracted_text = pytesseract.image_to_string(Image.open(image_path), config='--psm 6').strip()
        else:
            extracted_text = pytesseract.image_to_string(Image.open(image_path)).strip()
        print(f"Extracted text: {extracted_text}")
        # Handle specific edge cases
        if extracted_text.lower() in ['ia)', '11']:
            extracted_text = '11'
        elif extracted_text == '':
            print("No text found. Returning default value.")
            extracted_text = "Mar"  # Default fallback text
        
        return extracted_text

    except Exception as e:
        print(f"Error during OCR: {e}")
        return "Error"

def enhanced_image_to_string(image_path, number=True):
    """
    Converts text in an image to a string using OCR.

    Args:
        image_path (str): Path to the image file.
        number (bool): Whether to optimize OCR for numbers.

    Returns:
        str: Cleaned and processed extracted text from the image.
    """
    print(f"Extracting text from image: {image_path}")

    try:
        # Read the image
        img = cv2.imread(image_path)

        # Convert to grayscale
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # Resize to make the text more legible
        scale_factor = 3
        gray = cv2.resize(gray, None, fx=scale_factor, fy=scale_factor, interpolation=cv2.INTER_CUBIC)

        # Apply thresholding to improve contrast
        _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        # Optionally invert the image if the text is white on black
        thresh = cv2.bitwise_not(thresh)

        # Save the processed image for debugging (optional)
        # cv2.imwrite("processed_image.png", thresh)

        # Perform OCR
        config = '--psm 6 -c tessedit_char_whitelist=0123456789:'
        extracted_text = pytesseract.image_to_string(thresh, config=config if number else '--psm 6').strip()

        print(f"Extracted text: {extracted_text}")

        # Handle specific edge cases
        if extracted_text.lower() in ['ia)', '11']:
            extracted_text = '11'
        elif extracted_text == '':
            print("No text found. Returning default value.")
            extracted_text = "Mar"  # Default fallback text

        return extracted_text

    except Exception as e:
        print(f"Error during OCR: {e}")
        return "Error"
    
def rnd_value(x,range=6):
    """
    Generates a random value within ±5 of the given number x.

    Args:
        x (int or float): The base value.

    Returns:
        int or float: A random value within the range [x - range, x + range].
    """
    return random.randint(x - range, x + range)

def print_running_apps(d):
    print(d.app_list_running())


def return_code_inst(text, target_string):
    """
    Finds the first line in a text containing a specific substring and a 6-digit number.

    Args:
        text (str): The multi-line text to search.
        target_string (str): The substring to look for.

    Returns:
        str or None: The first matching line, or None if no match is found.
    """
    for line in text.splitlines():
        # Check if the line contains the target string and a 6-digit number
        if target_string in line and re.search(r'\b\d{6}\b', line):
            return line[-6:]
    return None


def format_with_leading_zero(number: int) -> str:
    """
    Converts an integer to a string and ensures it has at least two digits.
    Adds a leading zero if the number has only one digit.

    :param number: An integer to format
    :return: A string with at least two digits
    """
    # Convert the number to a string
    num_str = str(number)
    
    # Check if it has one digit and prepend '0' if necessary
    if len(num_str) == 1:
        num_str = '0' + num_str

    return num_str

def strip_newlines_and_spaces(input_string):
    """
    Removes all newline characters (\n) and spaces from the input string.

    Args:
        input_string (str): The string to process.

    Returns:
        str: The processed string without newline characters and spaces.
    """
    return input_string.replace(" ", "").replace("\n\n", "/").replace("\n", "/")

def transform_list(months, month_dict):
    """
    Transforms each object in the input list using a mapping dictionary.

    Args:
        input_list (list): The list of objects to transform.
        mapping_dict (dict): The dictionary for mapping values.

    Returns:
        list: A new list with transformed objects.
    """
    return convert_strings_to_ints([month_dict[month] for month in months if month in month_dict])
    
def convert_strings_to_ints(string_list):
    """
    Converts a list of strings to a list of integers, removing any items that cannot be converted.

    Args:
        string_list (list): List of strings to convert.

    Returns:
        list: A list of integers, excluding invalid strings.
    """
    return [int(item) for item in string_list if is_convertible_to_int(item)]

def is_convertible_to_int(value):
    """
    Checks if a value can be converted to an integer.

    Args:
        value: The value to check.

    Returns:
        bool: True if the value can be converted to an integer, False otherwise.
    """
    try:
        int(value)
        return True
    except ValueError:
        return False


def is_region_completely_white(image_path, region):
    """
    Check if a specific region of an image is completely white.

    Args:
        image_path (str): Path to the image file.
        region (tuple): Region to check in the format (left, top, right, bottom).

    Returns:
        bool: True if the region is completely white, False otherwise.
    """
    try:
        return np.all(np.array(Image.open(image_path).convert("RGB").crop(region)) == [255, 255, 255])

    except Exception as e:
        print(f"Error: {e}")
        return False
    

def start_random_function(functions,d):
    """
    Selects and starts a random function from a list of functions.

    Args:
        functions (list): A list of callable functions.

    Returns:
        Any: The result of the executed function.
    """
    if not functions:
        raise ValueError("The function list is empty.")
    sleep(rnd_value(10))
    return random.choice(functions)(d)


def arch_swipe(d, start_x_range, start_y_range, end_x_delta_range, end_y_delta_range, steps,duration):
    """
    Perform an arch-shaped swipe using a quadratic Bézier curve.

    Parameters:
        d: Device instance for swipe.
        start_x_range: Tuple (min, max) for the start X-coordinate.
        start_y_range: Tuple (min, max) for the start Y-coordinate.
        end_x_delta_range: Tuple (min, max) for the delta X offset for the end point.
        end_y_delta_range: Tuple (min, max) for the delta Y offset for the end point.
        steps: Number of steps for the swipe (smoothness of the curve).
    """
    # Define start and end coordinates with randomness
    start_x = random.randint(*start_x_range)
    start_y = random.randint(*start_y_range)
    end_x = start_x + random.randint(*end_x_delta_range)
    end_y = start_y - random.randint(*end_y_delta_range)
    # Calculate the mid-point for the arch
    mid_x = (start_x + end_x) // 2 + random.randint(-50, 50)  # Add randomness to the arch
    mid_y = (start_y + end_y) // 2 - random.randint(50, 150)  # Curve upwards

    # Generate points along the arch (quadratic Bézier curve)
    path = []
    for t in range(steps + 1):
        t = t / steps  # Normalize t between 0 and 1
        x = int((1 - t) ** 2 * start_x + 2 * (1 - t) * t * mid_x + t ** 2 * end_x)
        y = int((1 - t) ** 2 * start_y + 2 * (1 - t) * t * mid_y + t ** 2 * end_y)
        path.append((x, y))

    # Swipe through the points with small delays
    for i in range(len(path) - 1):
        d.swipe(path[i][0], path[i][1], path[i + 1][0], path[i + 1][1], duration=duration)

        
def list_running_devices():
    """Lists all running Genymotion devices."""
    try:
        result = subprocess.run(
            [gmtoolPath, "admin", "list", "vms"],
            capture_output=True,
            text=True,
            check=True
        )
        devices = []
        print("Listing all running devices:")
        for line in result.stdout.splitlines():
            if "On" in line:  # Filter for running devices
                device_info = line.split("|")
                device_name = device_info[2].strip()  # The name should be in the third column
                print(f"Found running device: {device_name}")
                devices.append(device_name)
        return devices
    except subprocess.CalledProcessError as e:
        print(f"Error listing devices: {e.stderr}")
        return []
    

def get_device_name_by_ip(ip_address):
    """Returns the device name for the given IP address."""
    try:
        result = subprocess.run(
            [gmtoolPath, "admin", "list", "vms"],
            capture_output=True,
            text=True,
            check=True
        )
        # Parse the result to find the device with the matching IP
        for line in result.stdout.splitlines():
            # Split the line by '|' to extract the fields
            parts = line.split("|")
            if len(parts) >= 3:
                # Extract IP address and device name
                device_ip = parts[1].strip()  # IP is in the second column
                device_name = parts[2].strip()  # Device name is in the third column
                # Check if the IP matches
                if device_ip == ip_address:
                    return device_name
        # Return None if no matching device was found
        return None

    except subprocess.CalledProcessError as e:
        print(f"Error listing devices: {e.stderr}")
        return None
    

def restart_genymotion():
    """Restarts Genymotion on a Linux system."""
    try:
        print("Restarting Genymotion...")
        # Kill any running Genymotion processes except the players
        subprocess.run(["pkill", "-f", "genymotion"], check=False)
        print("Genymotion processes terminated.")
        time.sleep(3)  # Wait a few seconds to ensure processes are fully terminated
        
        # Start Genymotion in a background process using nohup
        subprocess.Popen(
            ["nohup", "bash", "-c", "cd /home/goldfish/Desktop/genymotion && env -u QT_QPA_PLATFORM_PLUGIN_PATH ./genymotion &"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            preexec_fn=os.setpgrp  # Prevents the process from being killed when the script exits
        )

        print("Genymotion restarted successfully.")
    except Exception as e:
        print(f"Error restarting Genymotion: {e}")


def restart_device(d):
    """Restarts a specific Genymotion device with timeouts and logs the process."""
    try:
        device_name = get_device_name_by_ip(d.serial)
        if device_name is None:
            logging.error(f"Device with IP {d.serial} not found.")
            return
        logging.info(f"Stopping device: {d.serial}")
        subprocess.run(["env", "-u", "QT_QPA_PLATFORM_PLUGIN_PATH", gmtoolPath, "admin", "stop", device_name], check=True)
        time.sleep(5)
        logging.info(f"Starting device: {d.serial}")
        for attempt in range(5):
            try:
                subprocess.run(["env", "-u", "QT_QPA_PLATFORM_PLUGIN_PATH", gmtoolPath, "admin", "start", device_name], check=True)
                logging.info(f"Device {d.serial} started successfully on attempt {attempt + 1}.")
                break
            except subprocess.CalledProcessError as e:
                logging.error(f"Attempt {attempt + 1} failed to start device {d.serial}: {e.stderr}")
                if attempt == 2:
                    logging.info(f"Attempting to restart Genymotion...")
                    restart_genymotion()
                if attempt == 4:
                    logging.error(f"Failed to start device {d.serial} after 5 attempts.")
                    raise
            time.sleep(10)
        logging.info(f"Waiting for device {d.serial} to boot...")
        time.sleep(30)
        logging.info(f"Device {d.serial} restarted successfully.")
    except subprocess.CalledProcessError as e:
        logging.error(f"Error restarting device {d.serial}: {e.stderr}")
    except subprocess.TimeoutExpired as e:
        logging.error(f"Timeout expired while restarting device {d.serial}: {e}")
    



# Use the device name or UUID from list_running_devices

def get_links_and_reasons_from_non_red_cells(file_path, sheet_name, link_column, reason_column):
    # Load workbook and sheet
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    twitter_report_dict = {key: i + 1 for i, key in enumerate(twitter_report_keys)}
    data = []
    # Iterate over the specified column
    for row in range(2, 300):  # Assuming the first row is a header
        cell_link = ws[f"{link_column}{row}"]
        cell_reason = ws[f"{reason_column}{row}"]

        # Check if the link cell background is NOT red
        cell_color = cell_link.fill.start_color.index
        if cell_color not in ["FFFF0000", "FF0000"]:  # Exclude red-colored cells
            if cell_link.value and isinstance(cell_link.value, str) and cell_link.value.startswith("http"):
                
                # If reason is already a number, use it directly
                if isinstance(cell_reason.value, (int, float)):
                    reason_number = int(cell_reason.value)
                else:
                    reason_text = cell_reason.value.strip() if isinstance(cell_reason.value, str) else ""
                    reason_number = twitter_report_dict.get(reason_text, 5)  # Default to 5 (Incitement)
                data.append((cell_link.value, reason_number))

    return data


def extract_number_pairs(s):
        return time_to_seconds(re.findall(r'\d+:\d+', s)[0])

def time_to_seconds(time_str):
    print(time_str)
    parts = time_str.split(":")
    if len(parts) != 2:
        raise ValueError("Input must be in 'minutes:seconds' format")
    
    minutes, seconds = parts
    
    try:
        minutes = int(minutes)
        seconds = int(seconds)
    except ValueError:
        raise ValueError("Both minutes and seconds should be integers.")
    
    return minutes * 60 + seconds


def start_and_close_app(d):

    # Start the app
    d.app_start("com.twitter.android")
    # sleep(5)

    # Close the app
    # d.app_stop("com.twitter.android")
    # sleep(2)

    print(f"App started and closed successfully on device {d.serial}.")
    
    
update_results_file("Likes")